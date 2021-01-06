import sys 
from PyQt5 import QtGui, QtCore
from PyQt5.QtGui import QColor, QPalette
from PyQt5.QtCore import QDate, QTime, QDateTime, Qt,QFile
from PyQt5.QtWidgets import QApplication, QWidget, QInputDialog, QLineEdit, QFileDialog, QMessageBox
from PyQt5.QtWidgets import* 
from PyQt5.uic import loadUi
import serial.tools.list_ports
from serial import SerialException
import serial
import struct
import os
import sys
import glob
from dataclasses import dataclass
import xml.etree.ElementTree as ET
import string
import resource
import time
from openpyxl.utils import get_column_letter
from openpyxl import Workbook
from openpyxl.styles import Alignment
from openpyxl.styles import Color, PatternFill, Font, Border, Side

from openpyxl.styles import colors
from openpyxl.cell import Cell

class MyForm(QMainWindow):
    def __init__(self):
        QMainWindow.__init__(self)
        loadUi("form.ui",self)
        
        #self.struct = list() 
        self.structs = list()
        self.macros = list()
        self.enums = list()
        self.unions = list()
        
        
        self.gen_Button.clicked.connect(self.generate_pack)
        self.load_xml_Button.clicked.connect(self.load_xml_file)
        self.excel_output_Button.clicked.connect(self.excel_generate_file)

        self.hfile_txt.setFont(QtGui.QFont('Arial', 15))
        self.cfile_txt.setFont(QtGui.QFont('Arial', 15))
        
        self.check_method = self.check_method_comboBox.currentText()
        self.endian = self.endian_comboBox.currentText()
        self.align = self.align_comboBox.currentText()
        self.pack_attr = self.pack_attr_comboBox.currentText()
       

    def update_config(self):
        self.check_method = self.check_method_comboBox.currentText()
        self.endian = self.endian_comboBox.currentText()
        self.align = self.align_comboBox.currentText()
        self.pack_attr = self.pack_attr_comboBox.currentText()
        
    def load_xml_file(self):
        options = QFileDialog.Options()
        options |= QFileDialog.DontUseNativeDialog
        
        self.app_file_location, _ = QFileDialog.getOpenFileName(self,"QFileDialog.getOpenFileName()", "","Python Files (*.xml);;All Files (*)", options=options)
        self.file_lineEdit.setText(self.app_file_location)
        try :
            if self.app_file_location:
               self.xml_tree = ET.parse(self.app_file_location)
               self.parse_xml()
               #QMessageBox.critical(self, 'Error', 'parse xml failed', QMessageBox.Yes | QMessageBox.No, QMessageBox.No)

        except:
            QMessageBox.critical(self, 'Error', 'parse xml failed', QMessageBox.Yes | QMessageBox.No, QMessageBox.No)

    def excel_generate_file(self):
  
        book = Workbook()
        sheet = book.active
        
        sheet.cell(row=1, column=1).value = "Tx"
        sheet.cell(row=1, column=2).value = "Frame name"
        sheet.cell(row=1, column=3).value = "Frame Length"
        sheet.cell(row=1, column=4).value = "Frame field"
        sheet.cell(row=1, column=5).value = "Description"
        sheet.cell(row=1, column=6).value = "Position"
        sheet.cell(row=1, column=7).value = "Default value"
        sheet.cell(row=1, column=8).value = "Unit"
        sheet.cell(row=1, column=9).value = "Comments"
        
        #adjustment
        #for rol in sheet.columns:
        #set width
        for i in  string.ascii_lowercase[:14]:          
            sheet.column_dimensions[i].width = 25
        #set text alignment    
        for i in range(10):
            currentCell = sheet.cell(row=1, column=i+1)
            currentCell.alignment = Alignment(horizontal='center')
        #set the color and borders of cells 
        thin = Side(border_style="thin", color="000000")
        for i in range(10):           
            currentCell = sheet.cell(row=1, column=i+1)
            currentCell.fill = PatternFill("solid", fgColor="00BBFF")
            currentCell.border = Border(top=thin, left=thin, right=thin, bottom=thin)

        book.save("example.xlsx")
        pass
           
           
    def parse_xml(self):   
        if self.xml_tree :
            root = self.xml_tree.getroot()
            #structs 
            
            self.structs.clear()
            count =0
            self.header_name = root.get('name')
            self.header_name = self.header_name.upper()
            
            
            for i in root: 

              if i.tag == "macro" :
                  macro = list()
                  for j in i:
                      macro.append( (j.get('name'), j.get('val'))   )
                  self.macros.append((i.get('name'),macro))

              elif  i.tag== "enum":
                  enum = list()
                  for j in i:
                      enum.append(  (j.get('name'), j.get('val'))      )
                  self.enums.append(  (i.get('name'),enum)  )
                  
              elif  i.tag ==  "union":
                  union = list()
                  for j in i:
                      
                      if j.tag == "struct":
                          struct_in_union = list()
                          for k in j : 
                              if k.get('type') !="":
                                  struct_in_union.append(  (k.get('type'), k.get('name'), k.get('bit') ) )
                      union.append( struct_in_union )
                  
                  self.unions.append( (i.get('name'),i.get('anonymous_type') ,union) )    
                  #print(self.unions)
                 
                    
              elif i.tag == "struct" :
                  struct = list()
                  for j in i:
                      struct.append( (j.get('type'), j.get('name')) )
                      #print( (j.get('type'), j.get('name')) )      
                  self.structs.append(  (i.get('name'),struct,i.get('encode'),i.get('config')) )
              
              

    
    
    def generate_pack(self):
        
        self.cfile_txt.setText("")
        self.hfile_txt.setText("")
        
        self.generate_cfile()
        self.generate_hfile()                       
             
    def generate_decode_handle(msg):
    
         return msg  
         
    def generate_encode_handle(msg):
        
         
         pass              
   
    def generate_cfile(self):
        
        self.update_config()
        
        if self.check_method == "crc8" or self.check_method == "sum8":
            crc_reserve_bytes = "1"
            crc_return_type = "uint8_t"
        elif self.check_method == "crc16" or self.check_method == "sum16":
            crc_reserve_bytes = "2"
            crc_return_type = "uint16_t"
        elif self.check_method == "crc32" or self.check_method == "sum32":
            crc_reserve_bytes = "4"
            crc_return_type = "uint32_t"
        
        msg = ""
        msg += ''
        
        msg += '#include \"%s.h\"\n' % (self.header_name.lower())  #include "xxx.h"
        
        ''' 
        for struct in self.structs:
            struct_name = struct[0]
            msg +=  '%s_t %s;\n' % (struct_name, struct_name)
        '''
        
        path = "source/%s.c" %(self.check_method)
        f_check_func = open(path, "r").read()
        
        
        #crc_function
        msg += f_check_func
        msg += "\n"
        
        # check function
        msg += "%s check_function(uint8_t* pack,uint16_t len){\n" \
                    "\t return %s(pack, len);\n"\
                "}\n\n" % ((crc_return_type),self.check_method)
        
        ## input argument
        for struct in self.structs:
            struct_name = struct[0]
            context = struct[1]
            struct_config = struct[3]
            arg = ""
            statement = "\t\n"
            if struct_config == "true":
                for i in context:
                    statement += "\t pack->%s = %s;\n" % (i[1],i[1])
                    if i == context[-1]:  # if the last 1
                        arg += "\n\t      %s %s" %(i[0],i[1])
                    else:
                        arg += "\n\t      %s %s," %(i[0],i[1])
                msg +="void %s_config(%s_t* pack,%s){\n" %(struct_name,struct_name,arg)
                msg += statement
                msg += "\n}\n"

        ## encode
 
        for struct in self.structs:
            
            struct_name = struct[0]
            struct_encode = struct[2]
            if struct_encode == "true":
               
                msg +="\n"
                '''
                msg += "#define PACK_%s_ID %s\n" %(struct_name,str(pack_id))
                msg += "#define PACK_%s_LEN sizeof(%s_t)\n"% (struct_name, struct_name)
                if self.check_method == "sum32" or self.check_method == "crc32":
                    msg += "#define PACK_CRC_LEN %s\n" % ("4")
                elif self.check_method == "sum16" or self.check_method == "crc16":
                    msg += "#define PACK_CRC_LEN %s\n" % ("2")
                elif self.check_method == "sum8" or self.check_method == "crc8":
                    msg += "#define PACK_CRC_LEN %s\n" % ("1")
                '''    
                msg += '''
uint16_t %(n)s_encode(uint8_t* data, %(n)s_t* pack, uint16_t len){
	pack->header.packet_type = PACK_%(n)s_ID;
	pack->header.len = sizeof(%(n)s_t)-1-1;
	pack->checksum = check_function(pack, (sizeof(%(n)s_t) -%(s)s));
	return sizeof(%(n)s_t);
}
            ''' %({'n': struct_name, 's': "4"})
        crc_value =''
        
        if self.check_method == "sum32" or self.check_method=="crc32":
            crc_value +='''
            
            crc_val =(uint32_t)(buf[ind+0]<<0)  + 
            						(uint32_t)(buf[ind+1]<<8)  + 
            						(uint32_t)(buf[ind+2]<<16) + 
            						(uint32_t)(buf[ind+3]<<24);
            '''
        elif self.check_method == "sum16"or self.check_method=="crc16":
            crc_value +='''
            
            crc_val =(uint16_t)(buf[ind+0]<<0)  + 
            						(uint16_t)(buf[ind+1]<<8);
            					
            '''
        elif self.check_method == "sum8"or self.check_method=="crc8":
            crc_value +='''
            
            crc_val =(uint8_t)(buf[ind+0]<<0);
				
            '''
        ##parser
        
        msg+= '''
uint8_t packet_parser(uint8_t* buf,uint8_t data,parse_state_t* ps){
	//need a struct to record FSM state
	uint32_t crc_val;
	buf[ps->rx_index] = data;
	
	switch (ps->state)
	{
		case PARSE_STATE_START:
			if (buf[ps->rx_index] == 0x55)
			{
				ps->state = PARSE_STATE_LEN;
			}
			break;
		case PARSE_STATE_LEN:
			if(buf[ps->rx_index]<250)   //packet should be smaller than 255 bytes
			{
				ps->data_len = buf[ps->rx_index];
				ps->state = PARSE_STATE_DATA;
			}
			break;
		case PARSE_STATE_DATA:
			ps->now_idx++;
			if(ps->now_idx == ps->data_len)
			{
				ps->state = PARSE_STATE_CHECK;
				uint8_t ind = (ps->data_len+2) - PACK_CRC_LEN;
				
                %s

				 if(%s(buf,ind) == crc_val){
					ps->rx_index = 0;
					ps->now_idx = 0;
                    ps->data_len = 0;
					ps->state = PARSE_STATE_START;
					return buf[2];
				 }

			}
			break;
	}
	ps->rx_index++;
	return 0;
}
         ''' % (crc_value,self.check_method)
        # weak deocde function
        msg+='''
            __attribute__ ((weak))
            uint8_t message_decode(uint8_t* buf,uint8_t msg_id){
        	switch (msg_id)
        	{
        		case 1:
        			
        			memset(buf,0,200);
        			break;
        			
        		case 2:
        			
        			memset(buf,0,200);
        			break;
        			
        		case 3:
        		
        			memset(buf,0,200);
        			break;
        
        	}
        }
        
        
        
        '''
        
        # send
        '''
        msg += "uint16_t uart_send(uint8_t *data,uint16_t len){\n" \
                    "\t "\
                "}\n\n" 
        '''
        
        ## IRQ_Handler
        
        ## main function
        
        
        
        
        self.cfile_txt.setText(msg)


    def generate_hfile(self):
        pack_attr_front = ""
        pack_attr_end  = ""
        self.update_config()
     
        #if self.pack_attr  == "pragma":
        #    pack_attr_front = "#pragma pack(push,%s)\n" %(self.align)
        #    pack_attr_end = "#pragma pack(pop)\n"
        
        
        
        msg = ""
        msg += "#include \"stdint.h\"  \n"
        #msg += "#pragma once"
        msg += "#ifndef __%s_H_ \n" %(self.header_name.upper())
        msg += "#define __%s_H_ \n" %(self.header_name.upper())
        
        
       
        # print macros
        for macro in self.macros :
            macro_name = macro[0]
            macro_vars = macro[1]
            msg += "/* "  +  macro_name + " macros" + " */ \n"
            
            for var in macro_vars :
                msg += "#define" + "\t" +  var[0] + "\t" + var[1]  + "\n" 
        
        msg += "\n"
        msg+= "/* "  +  macro_name + " enumeration" + " */ \n"
        # print enumeration
        for enum in self.enums:
            enum_name = enum[0]
            enum_vars = enum[1]
            
            msg+= "typedef enum "+ enum_name  +"{\n"
            for  var in enum_vars:
                msg += "\t" + var[0] + " = " +var[1] + ",\n"
            
            msg+= "\n}%s_e;\n" %(enum_name)
        
        msg += "\n"       
        
        msg += pack_attr_front
        
        # print unions 
        for union in self.unions:
            union_name = union[0]
            union_type = union[1]
            union_vars = union[2]
            msg += "\n"
            msg += "typedef union %s_union{\n\n" %(union_name)
            
            for structs in union_vars:
               
                msg += " struct {\n"
                for var in structs:
                    msg += "\t%s %s : %s;\n" %(var[0],var[1],var[2])
                msg += "\t};\n"
            msg += "\t%s %s_value;\n"%(union_type, union_name)
            msg +=  "\n}%s_u;\n"%( union_name)
                    
                    
        #print FSM state struct
        msg+="\n"
        msg+= '''
        
enum parser{
	PARSE_STATE_START=0,
	PARSE_STATE_LEN=1,
	PARSE_STATE_DATA=2,
	PARSE_STATE_CHECK =3
};
        
typedef struct parse_state{
	uint8_t now_idx;
	uint8_t rx_index;
	uint8_t state;
	uint8_t data_len;
}parse_state_t;
            '''
        msg+="\n"
        # print structs  //struct pack  little/big endian
        pack_id = 0
        for struct in self.structs:
            
            struct_name = struct[0]
            struct_vars = struct[1]
            struct_encode = struct[2]
            msg += "\n"
            if struct_encode == "true":
                pack_id+=1
                msg += "#define PACK_%s_ID %s\n" %(struct_name,str(pack_id))
                msg += "#define PACK_%s_LEN sizeof(%s_t)\n"% (struct_name, struct_name)
                if self.check_method == "sum32" or self.check_method == "crc32":
                    msg += "#define PACK_CRC_LEN %s\n" % ("4")
                elif self.check_method == "sum16" or self.check_method == "crc16":
                    msg += "#define PACK_CRC_LEN %s\n" % ("2")
                elif self.check_method == "sum8" or self.check_method == "crc8":
                    msg += "#define PACK_CRC_LEN %s\n" % ("1")
                           
            
            msg += "\n"            
            if self.pack_attr == "pragma":
                msg += "#pragma pack(%s)\n"%(self.align)
                msg += "#pragma scalar_storage_order %s-endian\n" %(self.endian)
            elif self.pack_attr == "attribute":
                msg += " __attribute__((packed, aligned(%s), scalar_storage_order(\"%s-endian\")))" %(self.align, self.endian)
            
            
            msg += "typedef struct "  + struct_name + "_struct{\n\n"
            for var in struct_vars:
                msg+="\t" + var[0] + " " + var[1]+ ";\n"
            msg += "\n}"+ struct_name +"_t;\n"
                
        msg += "\n"
        
        msg += pack_attr_end
        
        msg += "\n"
        
        ## funciton declaration
        #msg += "void"
        
        
        msg += "#endif\n"
        
        
        self.hfile_txt.setText(msg)




if __name__=="__main__":
    app = QApplication(sys.argv)
    ex=MyForm()
    ex.show()
    sys.exit(app.exec_())











